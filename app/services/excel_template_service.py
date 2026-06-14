from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from typing import ClassVar

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.db.models import GENDERS, PERSON_CATEGORIES, RELATIONSHIP_TYPES


@dataclass(frozen=True)
class ExcelTemplateFile:
    content: bytes
    filename: str


class ExcelTemplateService:
    filename: ClassVar[str] = "networking_import_template.xlsx"
    template_version: ClassVar[str] = "1.0.0"
    schema_version: ClassVar[str] = "1.0.0"

    people_headers: ClassVar[list[str]] = [
        "person_key",
        "first_name",
        "last_name",
        "middle_name",
        "nickname",
        "phone",
        "telegram_username",
        "birth_date",
        "birth_year_known",
        "gender",
        "category",
        "custom_category",
        "note",
        "how_met",
        "location",
        "workplace",
        "education_place",
    ]
    relationship_headers: ClassVar[list[str]] = [
        "from_person_key",
        "to_person_key",
        "relationship_type",
        "custom_label",
        "note",
        "is_bidirectional",
    ]
    children_headers: ClassVar[list[str]] = [
        "parent_person_key",
        "child_person_key",
        "parent_role",
        "note",
    ]
    import_error_headers: ClassVar[list[str]] = [
        "row_number",
        "sheet_name",
        "column_name",
        "error_code",
        "error_message",
        "suggested_fix",
    ]

    def create_template(self) -> ExcelTemplateFile:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        instructions = workbook.create_sheet("Instructions")
        people = workbook.create_sheet("People")
        relationships = workbook.create_sheet("Relationships")
        children = workbook.create_sheet("Children")
        categories = workbook.create_sheet("Categories")
        import_errors = workbook.create_sheet("Import_Errors")

        self._fill_instructions(instructions)
        self._fill_sheet_header(people, self.people_headers, required_columns={"person_key", "first_name"})
        self._fill_sheet_header(
            relationships,
            self.relationship_headers,
            required_columns={"from_person_key", "to_person_key", "relationship_type"},
        )
        self._fill_sheet_header(
            children,
            self.children_headers,
            required_columns={"parent_person_key", "child_person_key"},
        )
        self._fill_categories(categories)
        self._fill_sheet_header(import_errors, self.import_error_headers, required_columns=set())

        self._add_people_validations(people)
        self._add_relationship_validations(relationships)
        self._add_children_validations(children)

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            self._auto_width(worksheet)

        workbook.active = workbook.sheetnames.index("Instructions")

        output = BytesIO()
        workbook.save(output)

        return ExcelTemplateFile(
            content=output.getvalue(),
            filename=self.filename,
        )

    def _fill_instructions(self, worksheet) -> None:
        rows = [
            ("template_version", self.template_version),
            ("schema_version", self.schema_version),
            ("generated_at", datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")),
            ("required_columns", "People.person_key and People.first_name are required."),
            ("birth_date_formats", "Use YYYY-MM-DD when the year is known, or MM-DD when the year is unknown."),
            ("person_key", "Temporary unique key inside this workbook, for example p001 or ali_valiyev_01."),
            ("relationships", "from_person_key and to_person_key must exist in People.person_key."),
            ("children", "parent_person_key and child_person_key must exist in People.person_key."),
            ("formula_safety", "Text cells must not start with =, +, -, or @."),
            ("example_person", "person_key=p001, first_name=Ali, birth_date=1995-04-21, category=friend"),
            ("example_relationship", "from_person_key=p001, to_person_key=p002, relationship_type=friend"),
            ("example_child", "parent_person_key=p001, child_person_key=p003, parent_role=father"),
        ]

        worksheet.append(["key", "value"])

        for row in rows:
            worksheet.append(list(row))

        self._style_header(worksheet, required_columns=set())

    def _fill_sheet_header(
        self,
        worksheet,
        headers: list[str],
        required_columns: set[str],
    ) -> None:
        worksheet.append(headers)
        self._style_header(worksheet, required_columns=required_columns)

    def _fill_categories(self, worksheet) -> None:
        worksheet.append(["category", "relationship_type", "gender", "parent_role", "boolean_values"])

        max_rows = max(len(PERSON_CATEGORIES), len(RELATIONSHIP_TYPES), len(GENDERS), 3, 2)

        parent_roles = ["father", "mother", "parent"]
        boolean_values = ["true", "false"]

        for index in range(max_rows):
            worksheet.append(
                [
                    PERSON_CATEGORIES[index] if index < len(PERSON_CATEGORIES) else None,
                    RELATIONSHIP_TYPES[index] if index < len(RELATIONSHIP_TYPES) else None,
                    GENDERS[index] if index < len(GENDERS) else None,
                    parent_roles[index] if index < len(parent_roles) else None,
                    boolean_values[index] if index < len(boolean_values) else None,
                ],
            )

        self._style_header(worksheet, required_columns=set())

    def _add_people_validations(self, worksheet) -> None:
        self._add_list_validation(worksheet, "J", f"Categories!$C$2:$C${len(GENDERS) + 1}")
        self._add_list_validation(worksheet, "K", f"Categories!$A$2:$A${len(PERSON_CATEGORIES) + 1}")
        self._add_list_validation(worksheet, "I", "Categories!$E$2:$E$3")

    def _add_relationship_validations(self, worksheet) -> None:
        self._add_list_validation(worksheet, "C", f"Categories!$B$2:$B${len(RELATIONSHIP_TYPES) + 1}")
        self._add_list_validation(worksheet, "F", "Categories!$E$2:$E$3")

    def _add_children_validations(self, worksheet) -> None:
        self._add_list_validation(worksheet, "C", "Categories!$D$2:$D$4")

    @staticmethod
    def _add_list_validation(worksheet, column: str, source: str) -> None:
        validation = DataValidation(
            type="list",
            formula1=source,
            allow_blank=True,
        )
        worksheet.add_data_validation(validation)
        validation.add(f"{column}2:{column}5000")

    @staticmethod
    def _style_header(worksheet, required_columns: set[str]) -> None:
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

            if str(cell.value) in required_columns:
                cell.fill = yellow_fill

    @staticmethod
    def _auto_width(worksheet) -> None:
        for column_cells in worksheet.columns:
            max_length = 0

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 45)
