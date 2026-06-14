from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from typing import ClassVar

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.i18n import I18nService


@dataclass(frozen=True)
class ImportErrorItem:
    row_number: int | None
    sheet_name: str
    column_name: str | None
    error_code: str
    error_message: str
    suggested_fix: str | None = None


@dataclass(frozen=True)
class ExcelErrorReportFile:
    content: bytes
    filename: str


class ExcelErrorReportService:
    headers: ClassVar[list[str]] = [
        "row_number",
        "sheet_name",
        "column_name",
        "error_code",
        "error_message",
        "suggested_fix",
    ]

    def create_error_report(
        self,
        errors: list[ImportErrorItem],
        i18n: I18nService | None = None,
        lang: str = "uz",
    ) -> ExcelErrorReportFile:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Import_Errors"

        worksheet.append(self.headers)

        for error in errors:
            error_message = error.error_message

            if i18n is not None:
                error_message = i18n.t(
                    f"import.errors.{error.error_code}",
                    lang=lang,
                )

            worksheet.append(
                [
                    error.row_number,
                    error.sheet_name,
                    error.column_name,
                    error.error_code,
                    error_message,
                    error.suggested_fix,
                ],
            )

        self._style_header(worksheet)
        worksheet.freeze_panes = "A2"

        for column in worksheet.columns:
            column_letter = column[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 60)

        output = BytesIO()
        workbook.save(output)

        timestamp = datetime.now(UTC).strftime("%Y-%m-%d_%H-%M")

        return ExcelErrorReportFile(
            content=output.getvalue(),
            filename=f"networking_import_errors_{timestamp}.xlsx",
        )

    @staticmethod
    def _style_header(worksheet) -> None:
        fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
