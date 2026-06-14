from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from typing import Any, ClassVar

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import Person, Relationship
from app.services.people_service import PeopleService
from app.services.security_service import SecurityService


@dataclass(frozen=True)
class ExcelExportFile:
    content: bytes
    filename: str


class ExcelExportService:
    people_headers: ClassVar[list[str]] = [
        "person_key",
        "first_name",
        "last_name",
        "middle_name",
        "nickname",
        "phone",
        "telegram_username",
        "birth_date",
        "birth_year_known",
        "gender",
        "category",
        "custom_category",
        "note",
        "how_met",
        "location",
        "workplace",
        "education_place",
    ]
    relationship_headers: ClassVar[list[str]] = [
        "from_person_key",
        "to_person_key",
        "relationship_type",
        "custom_label",
        "note",
        "is_bidirectional",
    ]
    children_headers: ClassVar[list[str]] = [
        "parent_person_key",
        "child_person_key",
        "parent_role",
        "note",
    ]
    birthday_headers: ClassVar[list[str]] = [
        "person_key",
        "full_name",
        "birth_date",
        "birth_month",
        "birth_day",
        "birth_year_known",
    ]

    async def create_export(
        self,
        session: AsyncSession,
        user_id: int,
    ) -> ExcelExportFile | None:
        people = await self._get_active_people(session=session, user_id=user_id)

        if not people:
            return None

        relationships = await self._get_active_relationships(
            session=session,
            user_id=user_id,
            active_person_ids={person.id for person in people},
        )

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        people_sheet = workbook.create_sheet("People")
        relationships_sheet = workbook.create_sheet("Relationships")
        children_sheet = workbook.create_sheet("Children")
        birthdays_sheet = workbook.create_sheet("Birthdays")
        instructions_sheet = workbook.create_sheet("Instructions")

        person_key_by_id = {person.id: f"p{person.id}" for person in people}

        self._write_people_sheet(people_sheet, people, person_key_by_id)
        self._write_relationships_sheet(relationships_sheet, relationships, person_key_by_id)
        self._write_children_sheet(children_sheet, relationships, person_key_by_id)
        self._write_birthdays_sheet(birthdays_sheet, people, person_key_by_id)
        self._write_instructions_sheet(instructions_sheet)

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            self._style_header(worksheet)
            self._auto_width(worksheet)

        workbook.active = workbook.sheetnames.index("People")

        output = BytesIO()
        workbook.save(output)

        timestamp = datetime.now(UTC).strftime("%Y-%m-%d_%H-%M")

        return ExcelExportFile(
            content=output.getvalue(),
            filename=f"networking_export_{timestamp}.xlsx",
        )

    async def _get_active_people(
        self,
        session: AsyncSession,
        user_id: int,
    ) -> list[Person]:
        result = await session.execute(
            select(Person)
            .where(
                Person.user_id == user_id,
                Person.deleted_at.is_(None),
            )
            .order_by(Person.first_name.asc(), Person.last_name.asc(), Person.id.asc()),
        )

        return list(result.scalars().all())

    async def _get_active_relationships(
        self,
        session: AsyncSession,
        user_id: int,
        active_person_ids: set[int],
    ) -> list[Relationship]:
        if not active_person_ids:
            return []

        result = await session.execute(
            select(Relationship)
            .where(
                Relationship.user_id == user_id,
                Relationship.deleted_at.is_(None),
                Relationship.from_person_id.in_(active_person_ids),
                Relationship.to_person_id.in_(active_person_ids),
            )
            .order_by(Relationship.id.asc()),
        )

        return list(result.scalars().all())

    def _write_people_sheet(
        self,
        worksheet,
        people: list[Person],
        person_key_by_id: dict[int, str],
    ) -> None:
        worksheet.append(self.people_headers)

        for person in people:
            worksheet.append(
                [
                    person_key_by_id[person.id],
                    self.escape_text(person.first_name),
                    self.escape_text(person.last_name),
                    self.escape_text(person.middle_name),
                    self.escape_text(person.nickname),
                    self.escape_text(person.phone),
                    self.escape_text(person.telegram_username),
                    self.format_birth_date(person),
                    str(person.birth_year_known).lower(),
                    person.gender,
                    person.category,
                    self.escape_text(person.custom_category),
                    self.escape_text(person.note),
                    self.escape_text(person.how_met),
                    self.escape_text(person.location),
                    self.escape_text(person.workplace),
                    self.escape_text(person.education_place),
                ],
            )

    def _write_relationships_sheet(
        self,
        worksheet,
        relationships: list[Relationship],
        person_key_by_id: dict[int, str],
    ) -> None:
        worksheet.append(self.relationship_headers)

        for relationship in relationships:
            worksheet.append(
                [
                    person_key_by_id[relationship.from_person_id],
                    person_key_by_id[relationship.to_person_id],
                    relationship.relationship_type,
                    self.escape_text(relationship.custom_label),
                    self.escape_text(relationship.note),
                    str(relationship.is_bidirectional).lower(),
                ],
            )

    def _write_children_sheet(
        self,
        worksheet,
        relationships: list[Relationship],
        person_key_by_id: dict[int, str],
    ) -> None:
        worksheet.append(self.children_headers)

        for relationship in relationships:
            if relationship.relationship_type != "parent":
                continue

            worksheet.append(
                [
                    person_key_by_id[relationship.from_person_id],
                    person_key_by_id[relationship.to_person_id],
                    relationship.custom_label or "parent",
                    self.escape_text(relationship.note),
                ],
            )

    def _write_birthdays_sheet(
        self,
        worksheet,
        people: list[Person],
        person_key_by_id: dict[int, str],
    ) -> None:
        worksheet.append(self.birthday_headers)
        people_service = PeopleService()

        birthday_people = [person for person in people if person.birth_month and person.birth_day]

        birthday_people.sort(
            key=lambda person: (person.birth_month, person.birth_day, person.first_name.lower()),
        )

        for person in birthday_people:
            worksheet.append(
                [
                    person_key_by_id[person.id],
                    people_service.format_full_name(person),
                    self.format_birth_date(person),
                    person.birth_month,
                    person.birth_day,
                    str(person.birth_year_known).lower(),
                ],
            )

    def _write_instructions_sheet(self, worksheet) -> None:
        worksheet.append(["key", "value"])
        generated_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        worksheet.append(["generated_at", generated_at])
        worksheet.append(["active_only", "Only people and relationships with deleted_at IS NULL are exported."])
        worksheet.append(["formula_safety", "Values starting with =, +, -, or @ are escaped with apostrophe."])

    @staticmethod
    def format_birth_date(person: Person) -> str | None:
        if person.birth_year_known and person.birth_date:
            return person.birth_date.isoformat()

        if person.birth_month and person.birth_day:
            return f"{person.birth_month:02d}-{person.birth_day:02d}"

        return None

    @staticmethod
    def escape_text(value: Any) -> Any:
        if value is None:
            return None

        text = str(value)

        return SecurityService.escape_formula_like(text)

    @staticmethod
    def _style_header(worksheet) -> None:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    @staticmethod
    def _auto_width(worksheet) -> None:
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 50)
