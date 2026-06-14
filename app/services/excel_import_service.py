from __future__ import annotations

from dataclasses import dataclass, field
from datetime import UTC, datetime
from io import BytesIO
from typing import Any, ClassVar

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import GENDERS, PERSON_CATEGORIES, RELATIONSHIP_TYPES, ImportJob
from app.db.repositories.imports import ImportsRepository
from app.db.repositories.people import PeopleRepository
from app.services.excel_error_report_service import ImportErrorItem
from app.services.people_service import PeopleService, PeopleValidationError
from app.services.relationship_service import RelationshipService, RelationshipValidationError
from app.services.reminder_service import ReminderService

FORMULA_PREFIXES = ("=", "+", "-", "@")


@dataclass
class ParsedPersonRow:
    row_number: int
    person_key: str
    data: dict[str, Any]


@dataclass
class ParsedRelationshipRow:
    row_number: int
    from_person_key: str
    to_person_key: str
    relationship_type: str
    custom_label: str | None
    note: str | None
    is_bidirectional: bool | None


@dataclass
class ParsedChildRow:
    row_number: int
    parent_person_key: str
    child_person_key: str
    parent_role: str | None
    note: str | None


@dataclass
class ExcelImportPreview:
    people_count: int
    relationships_count: int
    children_count: int
    duplicate_count: int
    error_count: int
    import_mode: str = "Skip duplicates"


@dataclass
class ParsedExcelImport:
    people: list[ParsedPersonRow] = field(default_factory=list)
    relationships: list[ParsedRelationshipRow] = field(default_factory=list)
    children: list[ParsedChildRow] = field(default_factory=list)
    duplicate_person_keys: set[str] = field(default_factory=set)
    errors: list[ImportErrorItem] = field(default_factory=list)

    @property
    def preview(self) -> ExcelImportPreview:
        return ExcelImportPreview(
            people_count=len(self.people),
            relationships_count=len(self.relationships),
            children_count=len(self.children),
            duplicate_count=len(self.duplicate_person_keys),
            error_count=len(self.errors),
        )


@dataclass(frozen=True)
class ExcelImportResult:
    imported_people_count: int
    imported_relationships_count: int
    skipped_duplicates_count: int
    import_job_id: int


class ExcelImportValidationError(ValueError):
    def __init__(self, message_key: str, detail: str | None = None) -> None:
        self.message_key = message_key
        self.detail = detail
        super().__init__(message_key)


class ExcelImportService:
    MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024

    people_headers: ClassVar[list[str]] = [
        "person_key",
        "first_name",
        "last_name",
        "middle_name",
        "nickname",
        "phone",
        "telegram_username",
        "birth_date",
        "birth_year_known",
        "gender",
        "category",
        "custom_category",
        "note",
        "how_met",
        "location",
        "workplace",
        "education_place",
    ]
    relationship_headers: ClassVar[list[str]] = [
        "from_person_key",
        "to_person_key",
        "relationship_type",
        "custom_label",
        "note",
        "is_bidirectional",
    ]
    children_headers: ClassVar[list[str]] = [
        "parent_person_key",
        "child_person_key",
        "parent_role",
        "note",
    ]
    required_sheets: ClassVar[set[str]] = {"People", "Relationships", "Children"}

    def validate_document(
        self,
        filename: str,
        content_type: str | None,
        file_size: int,
    ) -> None:
        if not filename.lower().endswith(".xlsx"):
            raise ExcelImportValidationError("excel.invalid_extension")

        if file_size > self.MAX_FILE_SIZE_BYTES:
            raise ExcelImportValidationError("excel.file_too_large")

        allowed_content_types = {
            None,
            "",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        }

        if content_type not in allowed_content_types:
            raise ExcelImportValidationError("excel.invalid_extension")

    async def parse_file(
        self,
        session: AsyncSession,
        user_id: int,
        content: bytes,
    ) -> ParsedExcelImport:
        try:
            workbook = load_workbook(BytesIO(content), data_only=False)
        except Exception as exc:
            raise ExcelImportValidationError("excel.import_failed", detail=str(exc)) from exc

        parsed = ParsedExcelImport()

        self._validate_required_sheets(workbook=workbook, parsed=parsed)

        if parsed.errors:
            return parsed

        people_sheet = workbook["People"]
        relationships_sheet = workbook["Relationships"]
        children_sheet = workbook["Children"]

        people_header_map = self._validate_headers(
            worksheet=people_sheet,
            expected_headers=self.people_headers,
            parsed=parsed,
        )
        relationships_header_map = self._validate_headers(
            worksheet=relationships_sheet,
            expected_headers=self.relationship_headers,
            parsed=parsed,
        )
        children_header_map = self._validate_headers(
            worksheet=children_sheet,
            expected_headers=self.children_headers,
            parsed=parsed,
        )

        if parsed.errors:
            return parsed

        await self._parse_people(
            session=session,
            user_id=user_id,
            worksheet=people_sheet,
            header_map=people_header_map,
            parsed=parsed,
        )
        self._parse_relationships(
            worksheet=relationships_sheet,
            header_map=relationships_header_map,
            parsed=parsed,
        )
        self._parse_children(
            worksheet=children_sheet,
            header_map=children_header_map,
            parsed=parsed,
        )
        self._validate_relationship_references(parsed=parsed)
        self._validate_children_references(parsed=parsed)

        return parsed

    async def import_parsed(
        self,
        session: AsyncSession,
        user_id: int,
        filename: str,
        file_size: int,
        parsed: ParsedExcelImport,
    ) -> ExcelImportResult:
        if parsed.errors:
            raise ExcelImportValidationError("excel.import_failed")

        async with session.begin_nested():
            import_job = ImportJob(
                user_id=user_id,
                import_type="excel",
                filename=filename,
                file_size=file_size,
                status="importing",
                total_people=len(parsed.people),
                total_relationships=len(parsed.relationships) + len(parsed.children),
                total_errors=0,
            )
            session.add(import_job)
            await session.flush()

            people_repository = PeopleRepository(session)
            relationship_service = RelationshipService()
            reminder_service = ReminderService()

            person_by_key: dict[str, Any] = {}
            skipped_duplicates_count = 0
            imported_people_count = 0
            imported_relationships_count = 0

            for person_row in parsed.people:
                if person_row.person_key in parsed.duplicate_person_keys:
                    skipped_duplicates_count += 1
                    continue

                person = await people_repository.create_person(
                    user_id=user_id,
                    data=person_row.data,
                )
                person_by_key[person_row.person_key] = person
                imported_people_count += 1

                if person.birth_month and person.birth_day:
                    await reminder_service.ensure_default_birthday_reminders_for_person(
                        session=session,
                        user_id=user_id,
                        person=person,
                    )

            for relationship_row in parsed.relationships:
                if (
                    relationship_row.from_person_key not in person_by_key
                    or relationship_row.to_person_key not in person_by_key
                ):
                    continue

                from_person = person_by_key[relationship_row.from_person_key]
                to_person = person_by_key[relationship_row.to_person_key]

                try:
                    await relationship_service.create_relationship(
                        session=session,
                        user_id=user_id,
                        from_person_id=from_person.id,
                        to_person_id=to_person.id,
                        relationship_type=relationship_row.relationship_type,
                        custom_label=relationship_row.custom_label,
                        note=relationship_row.note,
                        is_bidirectional=relationship_row.is_bidirectional,
                    )
                    imported_relationships_count += 1
                except RelationshipValidationError:
                    continue

            for child_row in parsed.children:
                if child_row.parent_person_key not in person_by_key or child_row.child_person_key not in person_by_key:
                    continue

                parent = person_by_key[child_row.parent_person_key]
                child = person_by_key[child_row.child_person_key]

                try:
                    await relationship_service.create_relationship(
                        session=session,
                        user_id=user_id,
                        from_person_id=parent.id,
                        to_person_id=child.id,
                        relationship_type="parent",
                        custom_label=child_row.parent_role or "parent",
                        note=child_row.note,
                        is_bidirectional=False,
                        reverse_relationship_type="child",
                    )
                    imported_relationships_count += 1
                except RelationshipValidationError:
                    continue

            import_job.status = "completed"
            import_job.completed_at = datetime.now(UTC)
            import_job.total_people = imported_people_count
            import_job.total_relationships = imported_relationships_count
            import_job.total_errors = 0

            await session.flush()
            await session.refresh(import_job)

        return ExcelImportResult(
            imported_people_count=imported_people_count,
            imported_relationships_count=imported_relationships_count,
            skipped_duplicates_count=skipped_duplicates_count,
            import_job_id=import_job.id,
        )

    async def record_failed_import_job(
        self,
        session: AsyncSession,
        user_id: int,
        filename: str,
        file_size: int,
        errors: list[ImportErrorItem],
    ) -> ImportJob:
        import_job = ImportJob(
            user_id=user_id,
            import_type="excel",
            filename=filename,
            file_size=file_size,
            status="failed",
            total_people=0,
            total_relationships=0,
            total_errors=len(errors),
            completed_at=datetime.now(UTC),
            error_message="Excel validation failed.",
        )

        session.add(import_job)
        await session.flush()

        imports_repository = ImportsRepository(session)

        for error in errors:
            await imports_repository.add_error(
                import_job_id=import_job.id,
                error_code=error.error_code,
                error_message=error.error_message,
                sheet_name=error.sheet_name,
                row_number=error.row_number,
                column_name=error.column_name,
                suggested_fix=error.suggested_fix,
            )

        await session.refresh(import_job)

        return import_job

    def _validate_required_sheets(
        self,
        workbook,
        parsed: ParsedExcelImport,
    ) -> None:
        workbook_sheets = set(workbook.sheetnames)

        for sheet_name in sorted(self.required_sheets - workbook_sheets):
            parsed.errors.append(
                ImportErrorItem(
                    row_number=None,
                    sheet_name=sheet_name,
                    column_name=None,
                    error_code="REQUIRED_FIELD_MISSING",
                    error_message=f"Required sheet is missing: {sheet_name}",
                    suggested_fix=f"Add sheet named {sheet_name}.",
                ),
            )

    def _validate_headers(
        self,
        worksheet,
        expected_headers: list[str],
        parsed: ParsedExcelImport,
    ) -> dict[str, int]:
        actual_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]

        header_map = {header: index + 1 for index, header in enumerate(actual_headers) if header}

        for expected_header in expected_headers:
            if expected_header not in header_map:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=1,
                        sheet_name=worksheet.title,
                        column_name=expected_header,
                        error_code="REQUIRED_FIELD_MISSING",
                        error_message=f"Required header is missing: {expected_header}",
                        suggested_fix=f"Add column {expected_header}.",
                    ),
                )

        return header_map

    async def _parse_people(
        self,
        session: AsyncSession,
        user_id: int,
        worksheet,
        header_map: dict[str, int],
        parsed: ParsedExcelImport,
    ) -> None:
        seen_person_keys: set[str] = set()
        people_service = PeopleService()
        people_repository = PeopleRepository(session)

        for row_number in range(2, worksheet.max_row + 1):
            row = self._read_row(worksheet, header_map, row_number)

            if self._row_is_empty(row):
                continue

            self._check_formula_injection(
                worksheet=worksheet,
                header_map=header_map,
                row_number=row_number,
                parsed=parsed,
            )

            person_key = self._clean_text(row.get("person_key"))
            first_name = self._clean_text(row.get("first_name"))

            if not person_key:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="People",
                        column_name="person_key",
                        error_code="REQUIRED_FIELD_MISSING",
                        error_message="person_key is required.",
                        suggested_fix="Fill person_key with a unique value such as p001.",
                    ),
                )
                continue

            if person_key in seen_person_keys:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="People",
                        column_name="person_key",
                        error_code="DUPLICATE_PERSON_KEY",
                        error_message=f"Duplicate person_key: {person_key}",
                        suggested_fix="Make person_key unique inside the workbook.",
                    ),
                )
                continue

            seen_person_keys.add(person_key)

            if not first_name:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="People",
                        column_name="first_name",
                        error_code="REQUIRED_FIELD_MISSING",
                        error_message="first_name is required.",
                        suggested_fix="Fill first_name.",
                    ),
                )
                continue

            category = self._clean_text(row.get("category"))

            if category and category not in PERSON_CATEGORIES:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="People",
                        column_name="category",
                        error_code="UNKNOWN_CATEGORY",
                        error_message=f"Unknown category: {category}",
                        suggested_fix="Use one of allowed values from Categories sheet.",
                    ),
                )
                continue

            if category == "custom" and not self._clean_text(row.get("custom_category")):
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="People",
                        column_name="custom_category",
                        error_code="CUSTOM_CATEGORY_REQUIRED",
                        error_message="custom_category is required when category=custom.",
                        suggested_fix="Fill custom_category.",
                    ),
                )
                continue

            gender = self._clean_text(row.get("gender"))

            if gender and gender not in GENDERS:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="People",
                        column_name="gender",
                        error_code="UNKNOWN_CATEGORY",
                        error_message=f"Unknown gender: {gender}",
                        suggested_fix="Use one of allowed values from Categories sheet.",
                    ),
                )
                continue

            try:
                data = people_service.prepare_create_data(
                    {
                        "first_name": first_name,
                        "last_name": self._clean_text(row.get("last_name")),
                        "middle_name": self._clean_text(row.get("middle_name")),
                        "nickname": self._clean_text(row.get("nickname")),
                        "phone": self._clean_text(row.get("phone")),
                        "telegram_username": self._clean_text(row.get("telegram_username")),
                        "birth_date": self._clean_text(row.get("birth_date")),
                        "gender": gender,
                        "category": category,
                        "custom_category": self._clean_text(row.get("custom_category")),
                        "note": self._clean_text(row.get("note")),
                        "how_met": self._clean_text(row.get("how_met")),
                        "location": self._clean_text(row.get("location")),
                        "workplace": self._clean_text(row.get("workplace")),
                        "education_place": self._clean_text(row.get("education_place")),
                    },
                )
            except PeopleValidationError as exc:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="People",
                        column_name=exc.field,
                        error_code="INVALID_DATE_FORMAT" if exc.field == "birth_date" else "REQUIRED_FIELD_MISSING",
                        error_message=exc.message_key,
                        suggested_fix="Check the field value and allowed format.",
                    ),
                )
                continue

            duplicates = await people_repository.find_duplicates(
                user_id=user_id,
                phone=data.get("phone"),
                telegram_username=data.get("telegram_username"),
                first_name=data.get("first_name"),
                last_name=data.get("last_name"),
                birth_date=data.get("birth_date"),
                nickname=data.get("nickname"),
            )

            if duplicates:
                parsed.duplicate_person_keys.add(person_key)

            parsed.people.append(
                ParsedPersonRow(
                    row_number=row_number,
                    person_key=person_key,
                    data=data,
                ),
            )

    def _parse_relationships(
        self,
        worksheet,
        header_map: dict[str, int],
        parsed: ParsedExcelImport,
    ) -> None:
        for row_number in range(2, worksheet.max_row + 1):
            row = self._read_row(worksheet, header_map, row_number)

            if self._row_is_empty(row):
                continue

            self._check_formula_injection(
                worksheet=worksheet,
                header_map=header_map,
                row_number=row_number,
                parsed=parsed,
            )

            from_person_key = self._clean_text(row.get("from_person_key"))
            to_person_key = self._clean_text(row.get("to_person_key"))
            relationship_type = self._clean_text(row.get("relationship_type"))

            if not from_person_key or not to_person_key or not relationship_type:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="Relationships",
                        column_name=None,
                        error_code="REQUIRED_FIELD_MISSING",
                        error_message="from_person_key, to_person_key and relationship_type are required.",
                        suggested_fix="Fill required relationship columns.",
                    ),
                )
                continue

            if relationship_type not in RELATIONSHIP_TYPES:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="Relationships",
                        column_name="relationship_type",
                        error_code="INVALID_RELATIONSHIP_TYPE",
                        error_message=f"Invalid relationship_type: {relationship_type}",
                        suggested_fix="Use allowed relationship_type from Categories sheet.",
                    ),
                )
                continue

            custom_label = self._clean_text(row.get("custom_label"))

            if relationship_type == "custom" and not custom_label:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="Relationships",
                        column_name="custom_label",
                        error_code="CUSTOM_LABEL_REQUIRED",
                        error_message="custom_label is required when relationship_type=custom.",
                        suggested_fix="Fill custom_label.",
                    ),
                )
                continue

            parsed.relationships.append(
                ParsedRelationshipRow(
                    row_number=row_number,
                    from_person_key=from_person_key,
                    to_person_key=to_person_key,
                    relationship_type=relationship_type,
                    custom_label=custom_label,
                    note=self._clean_text(row.get("note")),
                    is_bidirectional=self._parse_optional_bool(row.get("is_bidirectional")),
                ),
            )

    def _parse_children(
        self,
        worksheet,
        header_map: dict[str, int],
        parsed: ParsedExcelImport,
    ) -> None:
        for row_number in range(2, worksheet.max_row + 1):
            row = self._read_row(worksheet, header_map, row_number)

            if self._row_is_empty(row):
                continue

            self._check_formula_injection(
                worksheet=worksheet,
                header_map=header_map,
                row_number=row_number,
                parsed=parsed,
            )

            parent_person_key = self._clean_text(row.get("parent_person_key"))
            child_person_key = self._clean_text(row.get("child_person_key"))
            parent_role = self._clean_text(row.get("parent_role"))

            if not parent_person_key or not child_person_key:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="Children",
                        column_name=None,
                        error_code="REQUIRED_FIELD_MISSING",
                        error_message="parent_person_key and child_person_key are required.",
                        suggested_fix="Fill parent_person_key and child_person_key.",
                    ),
                )
                continue

            if parent_person_key == child_person_key:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="Children",
                        column_name="child_person_key",
                        error_code="INVALID_CHILD_RELATIONSHIP",
                        error_message="Parent and child cannot be the same person.",
                        suggested_fix="Use different person keys.",
                    ),
                )
                continue

            if parent_role and parent_role not in {"father", "mother", "parent"}:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name="Children",
                        column_name="parent_role",
                        error_code="UNKNOWN_CATEGORY",
                        error_message=f"Invalid parent_role: {parent_role}",
                        suggested_fix="Use father, mother or parent.",
                    ),
                )
                continue

            parsed.children.append(
                ParsedChildRow(
                    row_number=row_number,
                    parent_person_key=parent_person_key,
                    child_person_key=child_person_key,
                    parent_role=parent_role or "parent",
                    note=self._clean_text(row.get("note")),
                ),
            )

    def _validate_relationship_references(self, parsed: ParsedExcelImport) -> None:
        person_keys = {person.person_key for person in parsed.people}

        for relationship in parsed.relationships:
            if relationship.from_person_key not in person_keys:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=relationship.row_number,
                        sheet_name="Relationships",
                        column_name="from_person_key",
                        error_code="PERSON_KEY_NOT_FOUND",
                        error_message=f"Person key not found: {relationship.from_person_key}",
                        suggested_fix="Add this key to People sheet.",
                    ),
                )

            if relationship.to_person_key not in person_keys:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=relationship.row_number,
                        sheet_name="Relationships",
                        column_name="to_person_key",
                        error_code="PERSON_KEY_NOT_FOUND",
                        error_message=f"Person key not found: {relationship.to_person_key}",
                        suggested_fix="Add this key to People sheet.",
                    ),
                )

    def _validate_children_references(self, parsed: ParsedExcelImport) -> None:
        person_keys = {person.person_key for person in parsed.people}

        for child in parsed.children:
            if child.parent_person_key not in person_keys:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=child.row_number,
                        sheet_name="Children",
                        column_name="parent_person_key",
                        error_code="PERSON_KEY_NOT_FOUND",
                        error_message=f"Person key not found: {child.parent_person_key}",
                        suggested_fix="Add this key to People sheet.",
                    ),
                )

            if child.child_person_key not in person_keys:
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=child.row_number,
                        sheet_name="Children",
                        column_name="child_person_key",
                        error_code="PERSON_KEY_NOT_FOUND",
                        error_message=f"Person key not found: {child.child_person_key}",
                        suggested_fix="Add this key to People sheet.",
                    ),
                )

    def _check_formula_injection(
        self,
        worksheet,
        header_map: dict[str, int],
        row_number: int,
        parsed: ParsedExcelImport,
    ) -> None:
        for header, column_index in header_map.items():
            cell = worksheet.cell(row=row_number, column=column_index)
            value = cell.value

            if isinstance(value, str) and value.strip().startswith(FORMULA_PREFIXES):
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name=worksheet.title,
                        column_name=header,
                        error_code="FORMULA_NOT_ALLOWED",
                        error_message=f"Formula-like text is not allowed in {header}.",
                        suggested_fix="Remove leading =, +, -, or @.",
                    ),
                )

            if cell.data_type == "f":
                parsed.errors.append(
                    ImportErrorItem(
                        row_number=row_number,
                        sheet_name=worksheet.title,
                        column_name=header,
                        error_code="FORMULA_NOT_ALLOWED",
                        error_message=f"Formula is not allowed in {header}.",
                        suggested_fix="Replace formula with plain text.",
                    ),
                )

    @staticmethod
    def _read_row(
        worksheet,
        header_map: dict[str, int],
        row_number: int,
    ) -> dict[str, Any]:
        return {
            header: worksheet.cell(row=row_number, column=column_index).value
            for header, column_index in header_map.items()
        }

    @staticmethod
    def _row_is_empty(row: dict[str, Any]) -> bool:
        return all(value is None or str(value).strip() == "" for value in row.values())

    @staticmethod
    def _clean_text(value: Any) -> str | None:
        if value is None:
            return None

        cleaned = str(value).strip()

        return cleaned or None

    @staticmethod
    def _parse_optional_bool(value: Any) -> bool | None:
        if value is None or str(value).strip() == "":
            return None

        normalized = str(value).strip().lower()

        if normalized in {"true", "1", "yes", "ha", "да"}:
            return True

        if normalized in {"false", "0", "no", "yo‘q", "yo'q", "нет"}:
            return False

        return None
