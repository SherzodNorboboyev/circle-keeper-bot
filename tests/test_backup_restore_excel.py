from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.repositories.people import PeopleRepository
from app.db.repositories.users import UserRepository
from app.services.backup_service import BackupService
from app.services.excel_export_service import ExcelExportService
from app.services.excel_import_service import ExcelImportService
from app.services.excel_template_service import ExcelTemplateService
from app.services.people_service import PeopleService
from app.services.restore_service import RestoreService


class DummyBot:
    async def send_document(self, *args, **kwargs):
        raise RuntimeError("not used in tests")


async def create_user(
    session: AsyncSession,
    telegram_user_id: int,
) -> int:
    user = await UserRepository(session).upsert_from_telegram(
        telegram_user_id=telegram_user_id,
        chat_id=telegram_user_id,
        username=f"user_{telegram_user_id}",
        first_name="Test",
        last_name="User",
        language_code="uz",
        is_admin=False,
        default_timezone="Asia/Tashkent",
    )
    await session.flush()
    return user.id


async def create_person(
    session: AsyncSession,
    user_id: int,
    first_name: str,
):
    data = PeopleService().prepare_create_data(
        {
            "first_name": first_name,
            "birth_date": "1995-04-21",
            "category": "friend",
        },
    )
    return await PeopleRepository(session).create_person(user_id=user_id, data=data)


def build_import_workbook(
    people_rows: list[list[object]],
    relationship_rows: list[list[object]] | None = None,
    children_rows: list[list[object]] | None = None,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    people = workbook.create_sheet("People")
    people.append(ExcelTemplateService.people_headers)
    for row in people_rows:
        people.append(row)

    relationships = workbook.create_sheet("Relationships")
    relationships.append(ExcelTemplateService.relationship_headers)
    for row in relationship_rows or []:
        relationships.append(row)

    children = workbook.create_sheet("Children")
    children.append(ExcelTemplateService.children_headers)
    for row in children_rows or []:
        children.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def test_json_backup_contains_only_one_user_data(sqlite_session: AsyncSession) -> None:
    first_user_id = await create_user(sqlite_session, telegram_user_id=51001)
    second_user_id = await create_user(sqlite_session, telegram_user_id=51002)

    await create_person(sqlite_session, first_user_id, "Ali")
    await create_person(sqlite_session, second_user_id, "Sardor")

    service = BackupService(session=sqlite_session, bot=DummyBot())
    payload = await service.generate_backup_payload(user_id=first_user_id)

    assert len(payload["people"]) == 1
    assert payload["people"][0]["first_name"] == "Ali"
    assert payload["user"]["telegram_user_id"] == 51001


def test_sha256_calculation() -> None:
    assert BackupService.calculate_sha256(b"abc") == "ba7816bf8f01cfea414140de5dae2223b00361a396177a9cb410ff61f20015ad"


async def test_restore_replace_all_does_not_touch_other_user(sqlite_session: AsyncSession) -> None:
    first_user_id = await create_user(sqlite_session, telegram_user_id=51003)
    second_user_id = await create_user(sqlite_session, telegram_user_id=51004)

    old_person = await create_person(sqlite_session, first_user_id, "Old")
    other_person = await create_person(sqlite_session, second_user_id, "Other")

    service = BackupService(session=sqlite_session, bot=DummyBot())
    payload = await service.generate_backup_payload(user_id=first_user_id)
    payload["people"][0]["first_name"] = "Restored"
    payload["metadata"]["sha256"] = BackupService.calculate_payload_checksum(payload)

    await RestoreService().replace_all(
        session=sqlite_session,
        user_id=first_user_id,
        payload=payload,
    )

    first_user_people = await PeopleRepository(sqlite_session).list_people(user_id=first_user_id)
    second_user_people = await PeopleRepository(sqlite_session).list_people(user_id=second_user_id)

    assert old_person.id != first_user_people[0].id
    assert first_user_people[0].first_name == "Restored"
    assert second_user_people[0].id == other_person.id
    assert second_user_people[0].first_name == "Other"


def test_excel_template_sheets_and_headers() -> None:
    template = ExcelTemplateService().create_template()
    workbook = load_workbook(BytesIO(template.content))

    assert set(["Instructions", "People", "Relationships", "Children", "Categories", "Import_Errors"]).issubset(
        set(workbook.sheetnames)
    )
    assert [cell.value for cell in workbook["People"][1]] == ExcelTemplateService.people_headers
    assert [cell.value for cell in workbook["Relationships"][1]] == ExcelTemplateService.relationship_headers
    assert [cell.value for cell in workbook["Children"][1]] == ExcelTemplateService.children_headers


async def test_excel_import_valid_file(sqlite_session: AsyncSession) -> None:
    user_id = await create_user(sqlite_session, telegram_user_id=51005)

    content = build_import_workbook(
        people_rows=[
            [
                "p001",
                "Ali",
                "Valiyev",
                None,
                "Alish",
                None,
                "ali",
                "1995-04-21",
                "true",
                "male",
                "friend",
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "p002",
                "Sardor",
                "Karimov",
                None,
                None,
                None,
                None,
                "04-22",
                "false",
                "male",
                "colleague",
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ],
        relationship_rows=[["p001", "p002", "friend", None, None, "true"]],
    )

    service = ExcelImportService()
    parsed = await service.parse_file(sqlite_session, user_id=user_id, content=content)

    assert parsed.errors == []
    assert parsed.preview.people_count == 2
    assert parsed.preview.relationships_count == 1

    result = await service.import_parsed(
        session=sqlite_session,
        user_id=user_id,
        filename="valid.xlsx",
        file_size=len(content),
        parsed=parsed,
    )

    assert result.imported_people_count == 2
    assert result.imported_relationships_count == 1


async def test_excel_import_duplicate_person_key(sqlite_session: AsyncSession) -> None:
    user_id = await create_user(sqlite_session, telegram_user_id=51006)

    content = build_import_workbook(
        people_rows=[
            [
                "p001",
                "Ali",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "friend",
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "p001",
                "Sardor",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "friend",
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ],
    )

    parsed = await ExcelImportService().parse_file(sqlite_session, user_id=user_id, content=content)

    assert any(error.error_code == "DUPLICATE_PERSON_KEY" for error in parsed.errors)


async def test_excel_import_invalid_birth_date(sqlite_session: AsyncSession) -> None:
    user_id = await create_user(sqlite_session, telegram_user_id=51007)

    content = build_import_workbook(
        people_rows=[
            [
                "p001",
                "Ali",
                None,
                None,
                None,
                None,
                None,
                "31.02.1995",
                None,
                None,
                "friend",
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        ],
    )

    parsed = await ExcelImportService().parse_file(sqlite_session, user_id=user_id, content=content)

    assert any(error.error_code == "INVALID_DATE_FORMAT" for error in parsed.errors)


async def test_excel_import_custom_category_required(sqlite_session: AsyncSession) -> None:
    user_id = await create_user(sqlite_session, telegram_user_id=51008)

    content = build_import_workbook(
        people_rows=[
            [
                "p001",
                "Ali",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "custom",
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        ],
    )

    parsed = await ExcelImportService().parse_file(sqlite_session, user_id=user_id, content=content)

    assert any(error.error_code == "CUSTOM_CATEGORY_REQUIRED" for error in parsed.errors)


async def test_excel_import_relationship_person_key_not_found(sqlite_session: AsyncSession) -> None:
    user_id = await create_user(sqlite_session, telegram_user_id=51009)

    content = build_import_workbook(
        people_rows=[
            [
                "p001",
                "Ali",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "friend",
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        ],
        relationship_rows=[["p001", "p404", "friend", None, None, "true"]],
    )

    parsed = await ExcelImportService().parse_file(sqlite_session, user_id=user_id, content=content)

    assert any(error.error_code == "PERSON_KEY_NOT_FOUND" for error in parsed.errors)


async def test_excel_import_formula_injection_reject(sqlite_session: AsyncSession) -> None:
    user_id = await create_user(sqlite_session, telegram_user_id=51010)

    content = build_import_workbook(
        people_rows=[
            [
                "p001",
                "=HACK",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "friend",
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        ],
    )

    parsed = await ExcelImportService().parse_file(sqlite_session, user_id=user_id, content=content)

    assert any(error.error_code == "FORMULA_NOT_ALLOWED" for error in parsed.errors)


async def test_excel_export_active_only(sqlite_session: AsyncSession) -> None:
    user_id = await create_user(sqlite_session, telegram_user_id=51011)
    await create_person(sqlite_session, user_id, "Active")
    deleted_person = await create_person(sqlite_session, user_id, "Deleted")

    await PeopleRepository(sqlite_session).soft_delete_person(user_id=user_id, person_id=deleted_person.id)

    export_file = await ExcelExportService().create_export(sqlite_session, user_id=user_id)

    assert export_file is not None

    workbook = load_workbook(BytesIO(export_file.content))
    people_sheet = workbook["People"]

    exported_names = [
        people_sheet.cell(row=row_number, column=2).value for row_number in range(2, people_sheet.max_row + 1)
    ]

    assert exported_names == ["Active"]
